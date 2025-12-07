"""
table_writer.py
═══════════════════════════════════════════════════════════════════════════════
Модуль для создания КРАСИВЫХ и ЧИТАЕМЫХ Excel таблиц на основе данных из LLM.

✅ АДАПТИРОВАНО под ваш стиль excel_writer.py!
✅ Сохранение + пост-форматирование через load_workbook!
"""

import pandas as pd
import json
from typing import List, Dict, Any, Optional
from pathlib import Path
import datetime
import logging
from openpyxl import load_workbook
from openpyxl.styles import (
    Font, PatternFill, Alignment, Border, Side
)
from openpyxl.utils import get_column_letter


from settings import DEFAULT_COLUMNS, OUTPUT_DIR
from logger_cfg import setup_logger


logger = setup_logger(__name__)


# ════════════════════════════════════════════════════════════════════════════
# БЕЗОПАСНЫЕ УТИЛИТЫ (БЕЗ ИЗМЕНЕНИЙ)
# ════════════════════════════════════════════════════════════════════════════
def safe_str(value: Any, default: str = '') -> str:
    if value is None: return default
    if isinstance(value, str): return value.strip()
    try: return str(value).strip()
    except: return default

def safe_dict(data: Any, default: Dict = {}) -> Dict:
    return data if isinstance(data, dict) else default

def safe_list(data: Any, default: List = []) -> List:
    return data if isinstance(data, list) else default


# ════════════════════════════════════════════════════════════════════════════
# СОЗДАНИЕ ПУСТОЙ ТАБЛИЦЫ (БЕЗ ИЗМЕНЕНИЙ)
# ════════════════════════════════════════════════════════════════════════════
def create_empty_dataframe(columns: Optional[List[str]] = None) -> pd.DataFrame:
    if not columns: columns = DEFAULT_COLUMNS
    return pd.DataFrame(columns=columns)


# ════════════════════════════════════════════════════════════════════════════
# ПРЕОБРАЗОВАНИЕ ДАННЫХ (БЕЗ ИЗМЕНЕНИЙ)
# ════════════════════════════════════════════════════════════════════════════
def flatten_data(data: Dict[str, Any], pdf_filename: str) -> Dict[str, Any]:
    if "error" in data:
        return {
            'Адрес, комплекс': f"ОШИБКА: {safe_str(data.get('error'))}",
            'PDF-источник': pdf_filename,
            'Статус': 'Ошибка обработки'
        }

    owner_data = safe_dict(data.get('owner'))
    rental_data = safe_dict(data.get('rental_data'))
    objects = safe_list(data.get('objects_on_land'))

    def safe_object(obj: Any) -> Dict:
        if not obj: return {}
        return {
            'cadastral_number': safe_str(obj.get('cadastral_number')),
            'description': safe_str(obj.get('description'))
        }
    
    safe_objects = [safe_object(obj) for obj in objects if obj]
    objects_str = "; ".join([
        f"{obj['cadastral_number']} ({obj['description']})"
        for obj in safe_objects if obj['cadastral_number']
    ]) if safe_objects else '-'

    cadastral_num = safe_str(data.get('cadastral_number'))
    cadastral_zu = cadastral_num

    flat_row = {
        'Адрес, комплекс': safe_str(data.get('address')),
        'Наименование здания': safe_str(data.get('literal')),
        'Литера / Строение': safe_str(data.get('literal')),
        'Кадастр. номер ЗУ': f"{safe_str(data.get('cadastral_quarter', '-'))}",
        'Кадастр. номер здания': cadastral_zu,
        '№ помещения': safe_str(data.get('room_number')),
        'Этаж': safe_str(data.get('floor')),
        'Площадь (м²)': safe_str(data.get('area')),
        'Предполагаемое назначение': safe_str(data.get('permitted_use')),
        'Статус': safe_str(data.get('status')),
        'Арендатор': safe_str(data.get('tenant')),
        'Подтверждение из PDF': 'Автоматически',
        'Примечания и расхождения': safe_str(data.get('notes')),
        'Собственник': safe_str(data.get('owner')),
        'Обременение (аренда)': f"{safe_str(rental_data.get('rent_type'))} до {safe_str(rental_data.get('period_end'), 'Бессрочно')}".strip() or '-',
        'PDF-источник': pdf_filename
    }

    logger.info(f"✅ Данные распарсены: {pdf_filename}")
    return flat_row


def create_rows_from_llm_data(data: Dict[str, Any], pdf_name: str, columns: List[str]) -> List[Dict[str, Any]]:
    rows = []
    try:
        main_data = None
        if 'data' in data and data['data']: main_data = data['data']
        elif data and not 'error' in data: main_data = data
        else:
            logger.warning(f"⚠️ Нет данных LLM для {pdf_name}: {data}")
            row = {'PDF-источник': pdf_name, 'Статус': 'Нет данных AI'}
            for col in columns:
                if col not in row: row[col] = ''
            rows.append(row)
            return rows
        
        logger.info(f"✅ LLM данные найдены: {list(main_data.keys())}")
        flat_data = flatten_data(main_data, pdf_name)
        
        row = {}
        for col in columns:
            if col == '№ п/п':
                row[col] = 0
                continue
            row[col] = flat_data.get(col, '')
        
        rows.append(row)
        logger.info(f"✅ Строка создана: {pdf_name}")
        
    except Exception as e:
        logger.error(f"❌ Ошибка строки {pdf_name}: {e}")
        row = {'PDF-источник': pdf_name, 'Статус': f'Ошибка: {safe_str(e)}'}
        for col in columns:
            if col not in row: row[col] = ''
        rows.append(row)
    
    return rows


# ════════════════════════════════════════════════════════════════════════════
# 🔥 АДАПТИРОВАННОЕ ФОРМАТИРОВАНИЕ (как в вашем excel_writer.py) 🔥
# ════════════════════════════════════════════════════════════════════════════
def format_excel_file(excel_path: Path) -> bool:
    """
    ✅ ТОЧНАЯ КОПИЯ вашего форматирования из excel_writer.py!
    """
    try:
        logger.debug("🎨 Применение профессионального форматирования...")
        
        # Открываем файл
        wb = load_workbook(excel_path)
        ws = wb.active
        
        # ════════════════════════════════════════════════════════════════════
        # 1. ФОРМАТИРОВАНИЕ ШАПКИ (первая строка)
        # ════════════════════════════════════════════════════════════════════
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(name='Arial', size=11, bold=True, color="FFFFFF")
        header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
        
        logger.debug("✓ Шапка отформатирована")
        
        # ════════════════════════════════════════════════════════════════════
        # 2. АВТОШИРИНА КОЛОНОК
        # ════════════════════════════════════════════════════════════════════
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                try:
                    if cell.value:
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        logger.debug("✓ Ширина колонок установлена")
        
        # ════════════════════════════════════════════════════════════════════
        # 3. ФОРМАТИРОВАНИЕ ДАННЫХ + УСЛОВНОЕ ОКРАШИВАНИЕ
        # ════════════════════════════════════════════════════════════════════
        data_font = Font(name='Arial', size=10)
        data_alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # ✅ УСЛОВНОЕ ОКРАШИВАНИЕ ПО СОДЕРЖИМОМУ
        error_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
        success_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
        warning_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
        data_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        
        for row_idx, row in enumerate(ws.iter_rows(min_row=2), start=2):
            for cell in row:
                cell.font = data_font
                cell.alignment = data_alignment
                cell.border = thin_border
                
                # ✅ ЧЕРЕДОВАНИЕ + УСЛОВНОЕ
                cell_value = str(cell.value or '').lower()
                if row_idx % 2 == 0:
                    cell.fill = data_fill
                
                # ✅ КРАСИМ по содержимому!
                if 'ошиб' in cell_value or 'ошибка' in cell_value:
                    cell.fill = error_fill
                elif 'успех' in cell_value or 'готов' in cell_value:
                    cell.fill = success_fill
                elif '-' in cell_value or 'нет' in cell_value or 'н/д' in cell_value:
                    cell.fill = warning_fill
        
        # Границы для шапки
        for cell in ws[1]:
            cell.border = thin_border
        
        logger.debug("✓ Данные отформатированы с условным окрашиванием")
        
        # ════════════════════════════════════════════════════════════════════
        # 4. ВЫСОТА СТРОК
        # ════════════════════════════════════════════════════════════════════
        ws.row_dimensions[1].height = 30
        for row_idx in range(2, ws.max_row + 1):
            ws.row_dimensions[row_idx].height = 20
        
        logger.debug("✓ Высота строк установлена")
        
        # ════════════════════════════════════════════════════════════════════
        # 5. ЗАМОРОЗИТЬ ШАПКУ
        # ════════════════════════════════════════════════════════════════════
        ws.freeze_panes = "A2"
        logger.debug("✓ Шапка заморожена")
        
        # Сохраняем
        wb.save(excel_path)
        logger.debug("✓ Файл сохранен с форматированием")
        return True
        
    except Exception as e:
        logger.error(f"Ошибка при форматировании Excel: {type(e).__name__}: {str(e)}")
        return False


# ════════════════════════════════════════════════════════════════════════════
# 🔥 ОСНОВНАЯ ФУНКЦИЯ СОХРАНЕНИЯ (АДАПТИРОВАНА) 🔥
# ════════════════════════════════════════════════════════════════════════════
def save_to_excel(df: pd.DataFrame, output_path: Path) -> Path:
    """🎨 Сохраняет + форматирует ТОЧНО как ваш excel_writer.py!"""
    try:
        # ✅ Номера п/п
        if '№ п/п' in df.columns:
            df['№ п/п'] = range(1, len(df) + 1)
        
        # ✅ СОХРАНЯЕМ как в вашем коде
        output_path.parent.mkdir(parents=True, exist_ok=True)
        df.to_excel(output_path, sheet_name='Результаты', index=False)
        logger.debug(f"✓ DataFrame сохранен ({len(df)} строк)")
        
        # ✅ ФОРМАТИРУЕМ как в вашем коде
        format_excel_file(output_path)
        
        logger.info(f"💾 Красивый Excel сохранён: {output_path} ({len(df)} строк)")
        return output_path
        
    except Exception as e:
        logger.error(f"❌ Ошибка Excel {output_path}: {e}")
        raise


# ════════════════════════════════════════════════════════════════════════════
# ГЛАВНАЯ ФУНКЦИЯ (БЕЗ ИЗМЕНЕНИЙ)
# ════════════════════════════════════════════════════════════════════════════
def process_files_to_excel(pdf_files: List[Path], columns: List[str]) -> Dict[str, Any]:
    logger.info(f"📊 Создание Excel из {len(pdf_files)} файлов...")
    
    df = create_empty_dataframe(columns)
    stats = {'success': 0, 'failed': 0}
    
    for pdf_file in pdf_files:
        try:
            rows = create_rows_from_llm_data({}, pdf_file.name, columns)
            df_new = pd.DataFrame(rows)
            df = pd.concat([df, df_new], ignore_index=True)
            stats['success'] += 1
        except Exception as e:
            logger.error(f"❌ Обработка {pdf_file.name}: {e}")
            stats['failed'] += 1
    
    timestamp = datetime.datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"EGRN_Result_{timestamp}.xlsx"
    excel_path = OUTPUT_DIR / excel_filename
    
    excel_path = save_to_excel(df, excel_path)
    
    return {
        'success': True,
        'excel_path': str(excel_path),
        'excel_filename': excel_filename,
        'stats': stats
    }


# ════════════════════════════════════════════════════════════════════════════
# ДОПОЛНИТЕЛЬНАЯ ФУНКЦИЯ (из вашего кода)
# ════════════════════════════════════════════════════════════════════════════
def get_file_size(file_path: str) -> str:
    """Размер файла в читаемом формате."""
    try:
        size_bytes = Path(file_path).stat().st_size
        for unit in ['B', 'KB', 'MB', 'GB']:
            if size_bytes < 1024:
                return f"{size_bytes:.1f} {unit}"
            size_bytes /= 1024
        return f"{size_bytes:.1f} TB"
    except Exception as e:
        logger.error(f"Ошибка размера файла: {e}")
        return "Unknown"


if __name__ == "__main__":
    """🧪 Тест."""
    print("🧪 Тестируем table_writer (ваш стиль)...")
    
    cols = DEFAULT_COLUMNS
    df = create_empty_dataframe(cols)
    
    test_data = {
        "data": {
            "address": "ул. Тестовая, 1",
            "area": "1000 м²",
            "owner": {"full_name": "Иванов И.И."},
            "rental_data": {"tenant": "ООО Ромашка"},
            "objects_on_land": [{"cadastral_number": "74:36:0000000:123"}]
        }
    }
    
    rows = create_rows_from_llm_data(test_data, "test.pdf", cols)
    df_test = pd.DataFrame(rows)
    df = pd.concat([df, df_test], ignore_index=True)
    
    out_file = OUTPUT_DIR / "table_writer_test.xlsx"
    save_to_excel(df, out_file)
    print(f"✅ Тест OK: {out_file.absolute()}")
    print(f"   Размер: {get_file_size(out_file)}")
